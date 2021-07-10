from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from util.ShowOptions import *
import os


class ManipulateExcel:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        if "Excels" in os.listdir("./"):
            pass
        else:
            print("File save location not found!")
            print("Creating 'Excels' folder as save location...")
            os.system("mkdir Excels")
        
        self.path = "Excels/"
        self.filename = input("Insert the filename of the excel (no extension): ") + ".xlsx"
        self.cabeçalho_arr = []

        if self.filename in os.listdir(self.path):
            self.wb = load_workbook(self.path + self.filename)
        else:
            self.wb = Workbook()
        self.ws = self.wb.active

    def iniciar(self):
        self.AnalisarCabeçalho()
        os.system("cls")

        if self.cabeçalho_arr == []:
            self.CriarCabeçalho()
            os.system("cls")
        else:
            if input(f"\nThe following header was found: :\n{self.cabeçalho_arr}\nKeep? (Y/n): ").lower() == "n":
                self.EliminarCabeçalho()
                self.CriarCabeçalho()
        
        while True:
            os.system("cls")
            print(str(self.cabeçalho_arr) + "\n")
            print("\nSelect one of the option: ")
            printOptions(["Add Values", "Remove Values"])
            Add_Or_Remove = input("\nOption: ")
            
            if Add_Or_Remove == "0":
                os.system("cls")                
                self.AdicionarValores()
            elif Add_Or_Remove == "1":
                os.system("cls")
                self.RemoverValor()
            elif Add_Or_Remove.lower() == "s":
                break

        
        self.wb.save(self.path+self.filename)

    def CriarCabeçalho(self):
        cabeçalho_str = input("Enter the header elements separated by commas: ")
        self.cabeçalho_arr = [x.strip() for x in cabeçalho_str.split(",")]
        
        i = 0
        for col in range(1, len(self.cabeçalho_arr) + 1):
            self.ws[get_column_letter(col) + "1"].value = self.cabeçalho_arr[i]
            i += 1

    def AnalisarCabeçalho(self):
        isCellNone = False
        col = 1
        while not isCellNone:
            if self.ws[get_column_letter(col) + "1"].value != None:
                self.cabeçalho_arr.append(self.ws[get_column_letter(col) + "1"].value)
                col += 1
            else:
                break
    
    def EliminarCabeçalho(self):
        isCellNone = False
        col = 1
        while not isCellNone:
            if self.ws[get_column_letter(col) + "1"].value != None:
                self.ws[get_column_letter(col) + "1"].value = None
                col += 1
            else:
                self.cabeçalho_arr = []
                break
    
    
    def AnalisarValores(self, type = 0):
        arrayOptions = []
        arrayValores = []
        
        row = 2
        col = 1

        while True:
            if self.ws[get_column_letter(col) + str(row)].value != None:
                arrayValores.append(self.ws[get_column_letter(col) + str(row)].value)
                col += 1
            
            elif self.ws["A" + str(row + 1)].value != None:
                arrayOptions.append(arrayValores)
                arrayValores = []
                row +=1
                col = 1
            else:
                arrayOptions.append(arrayValores)
                break
        if type == 0:
            return arrayOptions
        else:
            for row in arrayOptions:
                print(str(row) + "\n")
    
    def RemoverValor(self):
        opts = self.AnalisarValores()
        
        while True:
            row = 2
            printOptions(opts)
            opt = input("Option: ")
            if  opt.lower() == "s":
                break
            try:
                opt = int(opt)
                row += opt
                self.ws.delete_rows(row)
                del opts[opt]
            except:
                print("Something went wrong...")
            os.system("cls")    

    def AdicionarValores(self):
        while True:
            if input("Enter new values? (Y/n)").lower() == "y":
                aRetornar = []
                for section in self.cabeçalho_arr:
                    aRetornar.append(input((f"{section}: ")))
                self.ws.append(aRetornar)
            else:
                break
            os.system("cls")
    
    

    


me = ManipulateExcel()
me.iniciar()
